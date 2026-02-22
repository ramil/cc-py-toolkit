#!/usr/bin/env python3
"""
SAP Payroll Implementation Validator v3
Validates Config Workbook and Migration File with fuzzy sheet name matching.
Handles naming variations: underscores, dashes, dots, spaces.
Includes v1.0 plugin checks for migration files and config workbook.
"""

import openpyxl
import json
import re
from collections import defaultdict

# ESSION anti-pattern list
ESSION_PATTERNS = [
    "MESSION", "PESSION", "PAESSION", "AESSION", "VESSION",
    "MOESSION", "KESSION", "TRESSION", "STESSION", "WESSION",
    "HESSION", "SAESSION", "BESSION", "ZESSION", "DESSION",
    "QUESSION", "UESSION", "FESSION", "NESSION", "SESSION"
]


def normalize_sheet_name(name):
    """Normalize sheet name for fuzzy matching: lowercase, strip separators"""
    return re.sub(r'[^a-z0-9]', '', name.lower())


def find_sheet(wb, keywords):
    """Find a sheet by keyword matching. Returns (sheet_name, worksheet) or (None, None)"""
    normalized_keywords = [k.lower() for k in keywords]
    for sname in wb.sheetnames:
        norm = normalize_sheet_name(sname)
        if all(k in norm for k in normalized_keywords):
            return sname, wb[sname]
    # Fallback: partial match on original name
    for sname in wb.sheetnames:
        slow = sname.lower()
        if all(k in slow for k in normalized_keywords):
            return sname, wb[sname]
    return None, None


def find_it_sheet(wb, it_code):
    """Find an infotype sheet (e.g., 'IT0001') regardless of suffix naming"""
    for sname in wb.sheetnames:
        if it_code in sname:
            return sname, wb[sname]
    return None, None


def count_data_rows(ws, start_row=2):
    """Count rows with at least one non-empty cell, starting from start_row"""
    count = 0
    for row in ws.iter_rows(min_row=start_row):
        if any(cell.value is not None for cell in row):
            count += 1
    return count


def get_all_text(ws):
    """Get all cell values as single uppercase string"""
    parts = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                parts.append(str(cell.value))
    return " ".join(parts).upper()


def check_ession(wb):
    """Check all sheets for ESSION hallucinated field names.
    Only flags actual SAP-like field names, not English words like 'professional'."""
    # Common English words that contain ESSION patterns (false positives)
    SAFE_WORDS = [
        "PROFESSIONAL", "PROFESSION", "CONFESSION", "OBSESSION", "REGRESSION",
        "COMPRESSION", "DEPRESSION", "EXPRESSION", "IMPRESSION", "PROGRESSION",
        "AGGRESSION", "DIGRESSION", "SUCCESSION", "RECESSION", "CONCESSION",
        "POSSESSION", "INTERCESSION", "SUPPRESSION", "ACCESSION", "PROCESSION",
        "DISCRETION", "SUBMISSION", "COMMISSION", "PERMISSION", "ADMISSION",
        "EMISSION", "TRANSMISSION", "SESSION", "MISSION", "PASSION",
        "ASSESSMENT", "REASSESSMENT"
    ]
    found = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value).upper()
                    # Skip if value contains safe English words
                    if any(sw in val for sw in SAFE_WORDS):
                        continue
                    # Skip long text (descriptions, not field names)
                    if len(str(cell.value)) > 30:
                        continue
                    for pat in ESSION_PATTERNS:
                        if pat in val and "ANTI" not in val and "BLOCK" not in val and "WARNING" not in val:
                            found.append(f"{sname}!{cell.coordinate}: {cell.value}")
    return found


def run_validation(config_file, migration_file, company_spec):
    """Full validation suite with fuzzy sheet matching"""
    # Set defaults for new fields
    if "benefits_approach" not in company_spec:
        company_spec["benefits_approach"] = "full"
    if "time_approach" not in company_spec:
        company_spec["time_approach"] = "full"

    results = {
        "company_id": company_spec.get("id"),
        "company_code": company_spec.get("code"),
        "company_name": company_spec.get("name"),
        "industry": company_spec.get("industry"),
        "config_workbook": {"checks": [], "issues": []},
        "migration_file": {"checks": [], "issues": []},
        "cross_validation": {"issues": []},
        "summary": {}
    }

    # Load workbooks
    try:
        cwb = openpyxl.load_workbook(config_file, data_only=True)
        mwb = openpyxl.load_workbook(migration_file, data_only=True)
    except Exception as e:
        results["error"] = str(e)
        results["summary"] = {"total_issues": 99, "validation_status": "ERROR", "score_pct": 0}
        return results

    cw_checks = results["config_workbook"]["checks"]
    cw_issues = results["config_workbook"]["issues"]
    mf_checks = results["migration_file"]["checks"]
    mf_issues = results["migration_file"]["issues"]
    xv_issues = results["cross_validation"]["issues"]

    # ============ CONFIG WORKBOOK CHECKS ============

    # CW-01: ESSION check
    ession = check_ession(cwb)
    cw_checks.append({"id": "CW-01", "name": "No ESSION fields", "pass": len(ession) == 0,
                       "detail": f"Found {len(ession)}: {ession[:3]}" if ession else "Clean"})

    # CW-02: Sheet count
    cw_checks.append({"id": "CW-02", "name": "Sheet count >= 15", "pass": len(cwb.sheetnames) >= 15,
                       "detail": f"{len(cwb.sheetnames)} sheets"})

    # CW-03: Key tabs exist and have data
    tab_checks = [
        ("CW-03a", "Enterprise Structure", [["enterprise", "structure"], ["enterprise"]], 10),
        ("CW-03b", "PSA Groupings", [["psa", "group"], ["psa"]], 5),
        ("CW-03c", "Feature Config", [["feature"]], 3),
        ("CW-03d", "Payroll Areas", [["payroll", "area"]], 1),
        ("CW-03e", "Payroll Calendar", [["payroll", "calendar"], ["calendar"]], 12),
        ("CW-03f", "Work Schedule", [["work", "schedule"], ["schedule"]], 3),
        ("CW-03g", "Wage Type Catalog", [["wage", "type"]], 20),
        ("CW-03h", "Processing/Eval Classes", [["process"]], 5),
        ("CW-03i", "WT Permissibility", [["permiss"]], 5),
        ("CW-03j", "Tax Authorities", [["tax", "author"], ["tax"]], 2),
        ("CW-03k", "Absence/Quota", [["absence"]], 3),
        ("CW-03l", "Symbolic Accounts", [["symbolic"]], 3),
        ("CW-03m", "House Bank", [["bank"]], 1),
        ("CW-03n", "Benefits", [["benefit"]], 1),
    ]

    found_tabs = {}
    for check_id, label, keyword_alternatives, min_rows in tab_checks:
        sname, ws = None, None
        for keywords in keyword_alternatives:
            sname, ws = find_sheet(cwb, keywords)
            if sname:
                break
        if sname and ws:
            rows = count_data_rows(ws)
            passed = rows >= min_rows
            cw_checks.append({"id": check_id, "name": f"{label} populated", "pass": passed,
                              "detail": f"'{sname}' has {rows} rows (need {min_rows})"})
            found_tabs[label] = (sname, ws)
            if not passed:
                cw_issues.append(f"{label}: only {rows} data rows (need {min_rows})")
        else:
            cw_checks.append({"id": check_id, "name": f"{label} exists", "pass": False,
                              "detail": "NOT FOUND"})
            cw_issues.append(f"{label} tab MISSING")

    # CW-04: All PSAs present
    if "Enterprise Structure" in found_tabs:
        _, ws = found_tabs["Enterprise Structure"]
        all_text = get_all_text(ws)
        missing = [p for p in company_spec.get("psas", []) if p.upper() not in all_text]
        cw_checks.append({"id": "CW-04", "name": "All PSAs in Enterprise", "pass": len(missing) == 0,
                          "detail": f"Missing: {missing}" if missing else f"All {len(company_spec.get('psas',[]))} found"})
        if missing:
            cw_issues.append(f"Missing PSAs: {missing}")

    # CW-05: Wage type count
    if "Wage Type Catalog" in found_tabs:
        _, ws = found_tabs["Wage Type Catalog"]
        wt_rows = count_data_rows(ws)
        expected = company_spec.get("wt_count", 35)
        pct = (wt_rows / expected * 100) if expected > 0 else 0
        passed = wt_rows >= expected * 0.7
        cw_checks.append({"id": "CW-05", "name": "Wage type completeness", "pass": passed,
                          "detail": f"{wt_rows} WTs vs {expected} expected ({pct:.0f}%)"})
        if not passed:
            cw_issues.append(f"Only {wt_rows} wage types, expected {expected}")
        results["config_workbook"]["wage_type_count"] = wt_rows

    # CW-06: SAP field names present
    sap_fields = ["MOLGA", "BUKRS", "WERKS", "BTRTL", "PERSG", "PERSK", "ABKRS", "LGART"]
    all_config_text = ""
    for sname in cwb.sheetnames[:8]:
        all_config_text += get_all_text(cwb[sname]) + " "
    found_fields = [f for f in sap_fields if f in all_config_text]
    cw_checks.append({"id": "CW-06", "name": "SAP field names present", "pass": len(found_fields) >= 5,
                      "detail": f"{len(found_fields)}/{len(sap_fields)}: {found_fields}"})

    # CW-07: Union config (if applicable)
    if company_spec.get("unions"):
        full_text = " ".join(get_all_text(cwb[s]) for s in cwb.sheetnames)
        has_union = any(x in full_text for x in ["UNION", "CBA", "DUES", "BARGAIN"])
        cw_checks.append({"id": "CW-07", "name": "Union config present", "pass": has_union,
                          "detail": "Found" if has_union else "MISSING for union company"})
        if not has_union:
            cw_issues.append("No union configuration for union company")

    # CW-08: No empty data tabs
    empty_tabs = []
    for sname in cwb.sheetnames:
        if any(x in sname.lower() for x in ["cover", "note", "readme", "toc", "index", "ai ", "analysis"]):
            continue
        ws = cwb[sname]
        if count_data_rows(ws) == 0:
            empty_tabs.append(sname)
    cw_checks.append({"id": "CW-08", "name": "No empty data tabs", "pass": len(empty_tabs) == 0,
                      "detail": f"Empty: {empty_tabs}" if empty_tabs else "All populated"})
    if empty_tabs:
        cw_issues.append(f"Empty tabs: {empty_tabs}")

    # CW-09: GL accounts in Symbolic Accounts — HKONT column must NOT be empty
    sym_sname, sym_ws = find_sheet(cwb, ["symbolic"])
    if sym_sname and sym_ws:
        # Find HKONT column (try various header names)
        hkont_col = None
        for row in sym_ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value and any(x in str(cell.value).upper() for x in ["HKONT", "GL ACCOUNT", "GL_ACCOUNT"]):
                    hkont_col = cell.column
                    break
            if hkont_col:
                break

        blank_glaccounts = []
        if hkont_col:
            for row in sym_ws.iter_rows(min_row=2):
                cell = row[hkont_col - 1] if len(row) >= hkont_col else None
                if cell and (cell.value is None or str(cell.value).strip() == ""):
                    blank_glaccounts.append(f"Row {cell.row}")

        passed = len(blank_glaccounts) == 0
        cw_checks.append({"id": "CW-09", "name": "GL accounts (HKONT) not blank", "pass": passed,
                          "detail": f"Blank: {blank_glaccounts[:5]}" if blank_glaccounts else "All HKONT populated"})
        if blank_glaccounts:
            cw_issues.append(f"Blank HKONT in Symbolic Accounts: {len(blank_glaccounts)} rows")
    else:
        cw_checks.append({"id": "CW-09", "name": "GL accounts (HKONT) not blank", "pass": False,
                          "detail": "Symbolic Accounts sheet not found"})

    # CW-10: Tax rates populated — SUI_ER_Rate or similar rate columns in Tax Authorities tab must not ALL be blank
    tax_sname, tax_ws = find_sheet(cwb, ["tax", "author"])
    if not tax_sname:
        tax_sname, tax_ws = find_sheet(cwb, ["tax"])
    if tax_sname and tax_ws:
        # Find rate columns (SUI_ER, SUI_EE, etc.)
        rate_cols = []
        for row in tax_ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value and any(x in str(cell.value).upper() for x in ["_RATE", "_ER", "_EE", "RATE"]):
                    rate_cols.append(cell.column)

        # Check if any rate column has data
        has_rates = False
        if rate_cols:
            for row in tax_ws.iter_rows(min_row=2):
                for col in rate_cols:
                    cell = row[col - 1] if len(row) >= col else None
                    if cell and cell.value is not None and str(cell.value).strip() != "":
                        has_rates = True
                        break
                if has_rates:
                    break

        cw_checks.append({"id": "CW-10", "name": "Tax rates populated", "pass": has_rates,
                          "detail": "Found tax rates" if has_rates else "No tax rates in Tax Authorities"})
        if not has_rates:
            cw_issues.append("No tax rates populated in Tax Authorities tab")
    else:
        cw_checks.append({"id": "CW-10", "name": "Tax rates populated", "pass": False,
                          "detail": "Tax Authorities sheet not found"})

    # ============ MIGRATION FILE CHECKS ============

    # MF-01: ESSION check
    ession_m = check_ession(mwb)
    mf_checks.append({"id": "MF-01", "name": "No ESSION fields", "pass": len(ession_m) == 0,
                       "detail": f"Found {len(ession_m)}" if ession_m else "Clean"})

    # MF-02: Sheet count
    mf_checks.append({"id": "MF-02", "name": "Sheet count >= 15", "pass": len(mwb.sheetnames) >= 15,
                       "detail": f"{len(mwb.sheetnames)} sheets"})

    # MF-03: Core infotype sheets
    core_its = ["IT0000", "IT0001", "IT0002", "IT0006", "IT0007", "IT0008", "IT0009", "IT0014"]
    found_core = []
    missing_core = []
    it_sheets = {}
    for it in core_its:
        sname, ws = find_it_sheet(mwb, it)
        if sname:
            found_core.append(it)
            it_sheets[it] = (sname, ws)
        else:
            missing_core.append(it)
    mf_checks.append({"id": "MF-03", "name": "Core infotypes present", "pass": len(missing_core) == 0,
                       "detail": f"Missing: {missing_core}" if missing_core else f"All {len(core_its)} found"})
    if missing_core:
        mf_issues.append(f"Missing core ITs: {missing_core}")

    # MF-04: Tax infotypes
    tax_its = ["IT0207", "IT0208", "IT0210"]
    for it in tax_its:
        sname, ws = find_it_sheet(mwb, it)
        if sname:
            it_sheets[it] = (sname, ws)
    found_tax = [it for it in tax_its if it in it_sheets]
    missing_tax = [it for it in tax_its if it not in it_sheets]
    mf_checks.append({"id": "MF-04", "name": "Tax infotypes present", "pass": len(missing_tax) == 0,
                       "detail": f"Missing: {missing_tax}" if missing_tax else "All 3 found"})

    # MF-05: Benefits infotypes (updated for v1.0 — 0 acceptable if benefits_approach="deductions_only")
    ben_its = ["IT0167", "IT0168", "IT0169", "IT0171"]
    for it in ben_its:
        sname, ws = find_it_sheet(mwb, it)
        if sname:
            it_sheets[it] = (sname, ws)
    found_ben = [it for it in ben_its if it in it_sheets]
    benefits_approach = company_spec.get("benefits_approach", "full")
    if benefits_approach == "deductions_only":
        passed_ben = True  # Accept 0 or more
        detail_ben = f"Found: {found_ben} (deductions_only approach)"
    else:
        passed_ben = len(found_ben) >= 2
        detail_ben = f"Found: {found_ben}"
    mf_checks.append({"id": "MF-05", "name": "Benefits infotypes", "pass": passed_ben,
                       "detail": detail_ben})

    # MF-06: Special infotypes
    for it in ["IT0194", "IT0559", "IT2006"]:
        sname, ws = find_it_sheet(mwb, it)
        if sname:
            it_sheets[it] = (sname, ws)

    if company_spec.get("garnishments"):
        has_garn = "IT0194" in it_sheets
        mf_checks.append({"id": "MF-06a", "name": "IT0194 Garnishment", "pass": has_garn,
                          "detail": "Found" if has_garn else "MISSING"})
    if company_spec.get("mid_year"):
        has_ytd = "IT0559" in it_sheets
        mf_checks.append({"id": "MF-06b", "name": "IT0559 YTD", "pass": has_ytd,
                          "detail": "Found" if has_ytd else "MISSING"})

    has_abs = "IT2006" in it_sheets
    mf_checks.append({"id": "MF-06c", "name": "IT2006 Absence Quotas", "pass": has_abs,
                      "detail": "Found" if has_abs else "MISSING"})

    # MF-07: Employee count
    ee_count = 0
    if "IT0001" in it_sheets:
        _, ws = it_sheets["IT0001"]
        ee_count = count_data_rows(ws)
    mf_checks.append({"id": "MF-07", "name": "Employee count >= 10", "pass": ee_count >= 10,
                       "detail": f"{ee_count} employees"})
    results["migration_file"]["employee_count"] = ee_count
    if ee_count < 10:
        mf_issues.append(f"Only {ee_count} employees (need 10+)")

    # MF-08: BSI geocode format — check TXJCD column (col D, index 4) only
    # Column E (TAXAUTH) legitimately contains "Federal" as a description
    bad_geocodes = []
    for it in ["IT0207", "IT0208"]:
        if it in it_sheets:
            _, ws = it_sheets[it]
            # Find TXJCD column (usually col 4)
            txjcd_col = 4  # default
            for row in ws.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    if cell.value and "TXJCD" in str(cell.value).upper():
                        txjcd_col = cell.column
                        break
            # Check TXJCD values — skip headers/labels
            for row in ws.iter_rows(min_row=4):  # skip title + subtitle + header
                cell = row[txjcd_col - 1] if len(row) >= txjcd_col else None
                if cell and cell.value:
                    val = str(cell.value).strip()
                    # Skip header-like values
                    if val.upper() in ["TXJCD", "TAX JURISDICTION", "GEOCODE", "TAX AREA"]:
                        continue
                    if val.upper() in ["FED", "FEDERAL", "STATE"] or (val and "-" not in val and len(val) > 2 and not any(c.isdigit() for c in val)):
                        bad_geocodes.append(f"{it}!{cell.coordinate}={val}")
    mf_checks.append({"id": "MF-08", "name": "No bad geocodes in TXJCD", "pass": len(bad_geocodes) == 0,
                       "detail": f"Bad: {bad_geocodes[:5]}" if bad_geocodes else "Clean BSI format"})
    if bad_geocodes:
        mf_issues.append(f"Bad geocodes: {bad_geocodes[:5]}")

    # MF-09: Federal geocode exists (00-000-0000)
    has_federal = False
    for it in ["IT0207", "IT0208"]:
        if it in it_sheets:
            _, ws = it_sheets[it]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.value is not None and str(cell.value).strip() == "00-000-0000":
                        has_federal = True
                        break
    mf_checks.append({"id": "MF-09", "name": "Federal geocode 00-000-0000", "pass": has_federal,
                       "detail": "Found" if has_federal else "MISSING"})

    # MF-10: IT0210 min 2 rows per employee
    if "IT0210" in it_sheets:
        _, ws = it_sheets["IT0210"]
        total_210 = count_data_rows(ws)
        avg_per_ee = (total_210 / ee_count) if ee_count > 0 else 0
        passed = avg_per_ee >= 1.8
        mf_checks.append({"id": "MF-10", "name": "IT0210 >= 2 rows/employee", "pass": passed,
                          "detail": f"{total_210} rows for {ee_count} EEs (avg {avg_per_ee:.1f})"})
        if not passed:
            mf_issues.append(f"IT0210: only {total_210} rows for {ee_count} employees")

    # MF-11: IT0014 sufficient deductions
    if "IT0014" in it_sheets:
        _, ws = it_sheets["IT0014"]
        total_14 = count_data_rows(ws)
        avg_ded = (total_14 / ee_count) if ee_count > 0 else 0
        passed = avg_ded >= 2
        mf_checks.append({"id": "MF-11", "name": "IT0014 sufficient deductions", "pass": passed,
                          "detail": f"{total_14} deductions for {ee_count} EEs (avg {avg_ded:.1f})"})

    # MF-12: 900-series SSNs
    if "IT0002" in it_sheets:
        _, ws = it_sheets["IT0002"]
        ssn_pattern = re.compile(r'^\d{3}-\d{2}-\d{4}$')
        ssns = []
        non_900 = []
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.value:
                    val = str(cell.value).strip()
                    if ssn_pattern.match(val):
                        ssns.append(val)
                        if not val.startswith("9"):
                            non_900.append(val)
        mf_checks.append({"id": "MF-12", "name": "900-series SSNs", "pass": len(non_900) == 0,
                          "detail": f"Non-900: {non_900[:3]}" if non_900 else f"{len(ssns)} SSNs all 900-series"})

    # MF-13: All states represented
    all_mig_text = " ".join(get_all_text(mwb[s]) for s in mwb.sheetnames)
    missing_states = [s for s in company_spec.get("states", []) if s in all_mig_text or s.upper() in all_mig_text]
    # Actually check they ARE there
    states_found = [s for s in company_spec.get("states", []) if s.upper() in all_mig_text]
    states_missing = [s for s in company_spec.get("states", []) if s.upper() not in all_mig_text]
    mf_checks.append({"id": "MF-13", "name": "All states represented", "pass": len(states_missing) == 0,
                       "detail": f"Missing: {states_missing}" if states_missing else f"All {len(company_spec.get('states',[]))} found"})

    # MF-14: Data Quality Review sheet
    dqr_name, dqr_ws = find_sheet(mwb, ["quality", "review"])
    if not dqr_name:
        dqr_name, dqr_ws = find_sheet(mwb, ["data", "quality"])
    mf_checks.append({"id": "MF-14", "name": "Data Quality Review sheet", "pass": dqr_name is not None,
                       "detail": f"Found: '{dqr_name}'" if dqr_name else "MISSING"})
    if not dqr_name:
        mf_issues.append("Missing Data Quality Review sheet")

    # MF-15: Cover sheet
    cover_name, _ = find_sheet(mwb, ["cover"])
    if not cover_name:
        cover_name, _ = find_sheet(mwb, ["summary"])
    mf_checks.append({"id": "MF-15", "name": "Cover/Summary sheet", "pass": cover_name is not None,
                       "detail": f"Found: '{cover_name}'" if cover_name else "MISSING"})

    # MF-16: No empty IT sheets
    empty_its = []
    for sname in mwb.sheetnames:
        if "IT" in sname and any(c.isdigit() for c in sname):
            ws = mwb[sname]
            if count_data_rows(ws) == 0:
                empty_its.append(sname)
    mf_checks.append({"id": "MF-16", "name": "No empty IT sheets", "pass": len(empty_its) == 0,
                       "detail": f"Empty: {empty_its}" if empty_its else "All populated"})
    if empty_its:
        mf_issues.append(f"Empty IT sheets: {empty_its}")

    # MF-17: Union employees (if applicable)
    if company_spec.get("unions"):
        has_union = any(x in all_mig_text for x in ["UNION", "CBA", "DUES"])
        mf_checks.append({"id": "MF-17", "name": "Union employee present", "pass": has_union,
                          "detail": "Found" if has_union else "MISSING"})

    # MF-18: All PAs in IT0001
    if "IT0001" in it_sheets:
        _, ws = it_sheets["IT0001"]
        it0001_text = get_all_text(ws)
        pa_codes = list(company_spec.get("pas", {}).keys())
        missing_pas = [p for p in pa_codes if p.upper() not in it0001_text]
        mf_checks.append({"id": "MF-18", "name": "All PAs in IT0001", "pass": len(missing_pas) == 0,
                          "detail": f"Missing: {missing_pas}" if missing_pas else f"All {len(pa_codes)} found"})

    # ============ NEW V1.0 MIGRATION FILE CHECKS ============

    # MF-19: IT0003 (Payroll Status) sheet exists and has data
    it0003_sname, it0003_ws = find_it_sheet(mwb, "IT0003")
    if it0003_sname:
        it0003_rows = count_data_rows(it0003_ws)
        passed_19 = it0003_rows > 0
        detail_19 = f"'{it0003_sname}' has {it0003_rows} rows"
    else:
        passed_19 = False
        detail_19 = "IT0003 sheet NOT FOUND"
    mf_checks.append({"id": "MF-19", "name": "IT0003 Payroll Status exists", "pass": passed_19,
                       "detail": detail_19})
    if not passed_19:
        mf_issues.append("IT0003 (Payroll Status) missing or empty")

    # MF-20: IT0041 (Date Specifications) sheet exists with >= 2 rows per employee
    it0041_sname, it0041_ws = find_it_sheet(mwb, "IT0041")
    if it0041_sname:
        it0041_rows = count_data_rows(it0041_ws)
        avg_0041 = (it0041_rows / ee_count) if ee_count > 0 else 0
        passed_20 = avg_0041 >= 1.8
        detail_20 = f"'{it0041_sname}' has {it0041_rows} rows ({avg_0041:.1f} per EE)"
    else:
        passed_20 = False
        detail_20 = "IT0041 sheet NOT FOUND"
    mf_checks.append({"id": "MF-20", "name": "IT0041 Date Specs >= 2/employee", "pass": passed_20,
                       "detail": detail_20})
    if not passed_20:
        mf_issues.append("IT0041 (Date Specifications) missing or insufficient rows per employee")

    # MF-21: IT0105 (Communication) sheet exists with >= 1 row per employee
    it0105_sname, it0105_ws = find_it_sheet(mwb, "IT0105")
    if it0105_sname:
        it0105_rows = count_data_rows(it0105_ws)
        avg_0105 = (it0105_rows / ee_count) if ee_count > 0 else 0
        passed_21 = avg_0105 >= 0.8
        detail_21 = f"'{it0105_sname}' has {it0105_rows} rows ({avg_0105:.1f} per EE)"
    else:
        passed_21 = False
        detail_21 = "IT0105 sheet NOT FOUND"
    mf_checks.append({"id": "MF-21", "name": "IT0105 Communication >= 1/employee", "pass": passed_21,
                       "detail": detail_21})
    if not passed_21:
        mf_issues.append("IT0105 (Communication) missing or insufficient rows per employee")

    # MF-22: IT0027 (Cost Distribution) exists IF company has concurrent_employment=True
    concurrent_employment = company_spec.get("concurrent_employment", False)
    if concurrent_employment:
        it0027_sname, it0027_ws = find_it_sheet(mwb, "IT0027")
        passed_22 = it0027_sname is not None
        detail_22 = f"Found '{it0027_sname}'" if passed_22 else "NOT FOUND (required for concurrent_employment)"
        mf_checks.append({"id": "MF-22", "name": "IT0027 Cost Dist (if concurrent_employment)", "pass": passed_22,
                           "detail": detail_22})
        if not passed_22:
            mf_issues.append("IT0027 (Cost Distribution) missing for concurrent_employment=True")

    # MF-23: IT0007 ZTEFN column consistency — if time_approach != "full", all ZTEFN values should be 9
    time_approach = company_spec.get("time_approach", "full")
    if time_approach != "full":
        it0007_sname, it0007_ws = find_it_sheet(mwb, "IT0007")
        if it0007_sname:
            # Find ZTEFN column
            ztefn_col = None
            for row in it0007_ws.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    if cell.value and "ZTEFN" in str(cell.value).upper():
                        ztefn_col = cell.column
                        break
                if ztefn_col:
                    break

            bad_ztefn = []
            if ztefn_col:
                for row in it0007_ws.iter_rows(min_row=2):
                    cell = row[ztefn_col - 1] if len(row) >= ztefn_col else None
                    if cell and cell.value is not None:
                        val = str(cell.value).strip()
                        if val and val != "9":
                            bad_ztefn.append(f"Row {cell.row}: {val}")

            passed_23 = len(bad_ztefn) == 0
            detail_23 = f"Non-9 values: {bad_ztefn[:5]}" if bad_ztefn else "All ZTEFN = 9"
            mf_checks.append({"id": "MF-23", "name": "IT0007 ZTEFN=9 (if time_approach!='full')", "pass": passed_23,
                               "detail": detail_23})
            if not passed_23:
                mf_issues.append(f"IT0007 ZTEFN inconsistency: {len(bad_ztefn)} non-9 values")
        else:
            mf_checks.append({"id": "MF-23", "name": "IT0007 ZTEFN=9 (if time_approach!='full')", "pass": False,
                               "detail": "IT0007 sheet NOT FOUND"})

    # ============ SUMMARY ============
    cw_pass = sum(1 for c in cw_checks if c["pass"])
    cw_fail = sum(1 for c in cw_checks if not c["pass"])
    mf_pass = sum(1 for c in mf_checks if c["pass"])
    mf_fail = sum(1 for c in mf_checks if not c["pass"])
    total_checks = len(cw_checks) + len(mf_checks)
    total_pass = cw_pass + mf_pass
    total_fail = cw_fail + mf_fail
    score = (total_pass / total_checks * 100) if total_checks > 0 else 0

    total_issues = len(cw_issues) + len(mf_issues) + len(xv_issues)
    if score >= 90:
        status = "PASS"
    elif score >= 70:
        status = "PASS WITH WARNINGS"
    else:
        status = "FAIL"

    results["config_workbook"]["sheet_count"] = len(cwb.sheetnames)
    results["config_workbook"]["sheets"] = cwb.sheetnames
    results["migration_file"]["sheet_count"] = len(mwb.sheetnames)
    results["migration_file"]["sheets"] = mwb.sheetnames

    results["summary"] = {
        "cw_pass": cw_pass, "cw_fail": cw_fail,
        "mf_pass": mf_pass, "mf_fail": mf_fail,
        "total_checks": total_checks, "total_pass": total_pass, "total_fail": total_fail,
        "score_pct": round(score, 1),
        "total_issues": total_issues,
        "validation_status": status
    }

    cwb.close()
    mwb.close()
    return results


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 4:
        print("Usage: python validator.py <config.xlsx> <migration.xlsx> <company_json>")
        sys.exit(1)
    config_path = sys.argv[1]
    migration_path = sys.argv[2]
    company = json.loads(sys.argv[3])
    results = run_validation(config_path, migration_path, company)
    print(json.dumps(results, indent=2))
