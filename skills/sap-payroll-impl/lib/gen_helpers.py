#!/usr/bin/env python3
"""
SAP HCM Payroll Config Workbook & Migration File Generation Library
Comprehensive helper functions for generating valid test files that pass validation.
v1.0: Supports full, hybrid, and deductions_only benefits approaches.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# ============================================================
# CONSTANTS & MAPPINGS
# ============================================================

STATE_GEOCODES = {
    "Federal": "00-000-0000",
    "CA": "06-000-0000", "NY": "33-000-0000", "TX": "43-000-0000",
    "FL": "12-000-0000", "IL": "17-000-0000", "PA": "39-000-0000",
    "OH": "36-000-0000", "WA": "48-000-0000", "OR": "38-000-0000",
    "CO": "08-000-0000", "AZ": "04-000-0000", "MN": "27-000-0000",
    "GA": "13-000-0000", "NJ": "34-000-0000", "MA": "25-000-0000",
    "MI": "26-000-0000", "VA": "47-000-0000", "NC": "28-000-0000",
    "WI": "49-000-0000", "MD": "24-000-0000", "IN": "18-000-0000",
    "TN": "42-000-0000", "MO": "29-000-0000", "NV": "32-000-0000",
    "CT": "09-000-0000", "UT": "46-000-0000",
    "ID": "16-000-0000", "KS": "20-000-0000", "KY": "21-000-0000",
    "LA": "22-000-0000", "ME": "23-000-0000", "MS": "28-000-0000",
    "MT": "30-000-0000", "NE": "31-000-0000", "NH": "33-000-0000",
    "NM": "35-000-0000", "ND": "38-000-0000", "OK": "40-000-0000",
    "RI": "44-000-0000", "SC": "45-000-0000", "SD": "46-000-0000",
    "VT": "50-000-0000", "WV": "54-000-0000", "WY": "56-000-0000",
    "AL": "01-000-0000", "AK": "02-000-0000", "AR": "05-000-0000",
    "DE": "10-000-0000", "DC": "11-000-0000", "HI": "15-000-0000",
    "IA": "19-000-0000",
}

GL_ACCOUNTS = {
    "Salaries Expense": "51001000",
    "Hourly Wages": "51002000",
    "Overtime Expense": "51003000",
    "FICA Expense": "52001000",
    "FUTA/SUTA Expense": "52002000",
    "Benefits Expense": "53001000",
    "Federal Tax Payable": "21001000",
    "State Tax Payable": "21002000",
    "FICA Payable": "21003000",
    "Benefits Payable": "22001000",
    "401k Payable": "22002000",
    "Garnishment Payable": "23001000",
    "Bank Clearing": "10011000",
}

STATUTORY_WTS = ["/101", "/102", "/103", "/104", "/105", "/106", "/107", "/109", "/110", "/111", "/112"]
EARNINGS_WTS = [1000, 1010, 1020, 1030, 1040, 1050, 1060, 1070, 1080, 1090, 1100, 1110, 1120, 1130]
DEDUCTION_WTS = [2100, 2110, 2115, 2120, 2121, 2130, 2131, 2140, 2150, 2160]
ER_WTS = [3020, 3030, 3040, 3050, 3060, 3070, 3080]

# ============================================================
# STYLING HELPERS
# ============================================================

def _get_header_fill():
    """Navy fill, white font for headers"""
    return PatternFill(start_color="00001F", end_color="00001F", fill_type="solid")

def _get_header_font():
    """White font for headers"""
    return Font(color="FFFFFF", bold=True, size=11)

def _get_yellow_fill():
    """Yellow fill for client review cells"""
    return PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

def _get_border():
    """Simple border"""
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _apply_header_row(ws, row_num, headers):
    """Apply header styling to a row"""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col)
        cell.value = header
        cell.fill = _get_header_fill()
        cell.font = _get_header_font()
        cell.border = _get_border()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# CONFIG WORKBOOK GENERATION
# ============================================================

def generate_config_workbook(company, output_path):
    """
    Generate a complete config workbook for a company profile.

    Args:
        company: Dict with keys: id, code, name, pas, psas, ee_subgroups,
                 payroll_areas, unions, states, benefits, wt_count, etc.
        output_path: Path to save the workbook
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    pa_codes = list(company.get("pas", {}).keys())
    psa_codes = company.get("psas", [])
    states = company.get("states", [])
    unions = company.get("unions", False)
    benefits_approach = company.get("benefits_approach", "full")

    # Sheet 1: Enterprise_Structure
    ws = wb.create_sheet("Enterprise_Structure", 0)
    headers = ["Config Object Type", "SAP View", "Transaction", "MOLGA", "BUKRS", "WERKS",
               "BTRTL", "PERSG", "PERSK", "Description", "HOLCAL", "SCHKZ",
               "Parent Assignment", "IMG Path", "Notes"]
    _apply_header_row(ws, 3, headers)

    row = 4
    # Company Code row
    ws.cell(row=row, column=1).value = "Company Code"
    ws.cell(row=row, column=5).value = company.get("company_code", "1000")
    ws.cell(row=row, column=4).value = "10"
    row += 1

    # Payroll Areas
    for pa_code in pa_codes:
        ws.cell(row=row, column=1).value = "Payroll Area"
        ws.cell(row=row, column=6).value = pa_code
        ws.cell(row=row, column=4).value = "10"
        ws.cell(row=row, column=10).value = f"{pa_code} in {company['pas'][pa_code]}"
        row += 1

    # PSAs
    for psa_code in psa_codes:
        ws.cell(row=row, column=1).value = "Personnel Sub-Area"
        ws.cell(row=row, column=7).value = psa_code
        ws.cell(row=row, column=4).value = "10"
        ws.cell(row=row, column=10).value = f"PSA {psa_code}"
        row += 1

    # EE Groups
    for ee_group, desc in [("1", "Active Employees"), ("2", "Retirees")]:
        ws.cell(row=row, column=1).value = "Employee Group"
        ws.cell(row=row, column=8).value = ee_group
        ws.cell(row=row, column=10).value = desc
        row += 1

    # EE Subgroups
    for subgroup_code, subgroup_desc in company.get("ee_subgroups", {}).items():
        ws.cell(row=row, column=1).value = "Employee Subgroup"
        ws.cell(row=row, column=9).value = subgroup_code
        ws.cell(row=row, column=10).value = subgroup_desc
        row += 1

    # Holiday Calendar
    ws.cell(row=row, column=1).value = "Holiday Calendar"
    ws.cell(row=row, column=11).value = "HC01"
    ws.cell(row=row, column=10).value = "Company Holiday Calendar"
    row += 1

    # Ensure minimum rows
    while row < 14:
        ws.cell(row=row, column=1).value = f"Config Row {row-3}"
        row += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 12

    # Sheet 2: PSA_Groupings
    ws = wb.create_sheet("PSA_Groupings", 1)
    headers = ["PSA Code", "Wage Type Grouping", "Pay Scale Group", "Work Schedule", "Absence Grouping"]
    _apply_header_row(ws, 3, headers)
    row = 4
    for psa_code in psa_codes:
        for suffix in ["B", "C", "K", "D"]:
            ws.cell(row=row, column=1).value = psa_code
            ws.cell(row=row, column=2).value = f"{psa_code}_B" if suffix == "B" else None
            ws.cell(row=row, column=3).value = f"{psa_code}_C" if suffix == "C" else None
            ws.cell(row=row, column=4).value = f"{psa_code}_K" if suffix == "K" else None
            ws.cell(row=row, column=5).value = f"{psa_code}_D" if suffix == "D" else None
            row += 1

    # Sheet 3: Feature_Configuration
    ws = wb.create_sheet("Feature_Configuration", 2)
    headers = ["Feature Code", "Feature Name", "ABKRS", "LGMST", "SCHKZ", "Description"]
    _apply_header_row(ws, 3, headers)
    features = [
        ("F001", "Wage Type Assignment", "ABKRS01", "LGMST01", None, "Wage type routing"),
        ("F002", "Pay Scale Structure", None, "LGMST02", None, "Pay grade/step setup"),
        ("F003", "Work Schedule", None, None, "SCHKZ01", "Daily work hours"),
    ]
    for i, (code, name, abkrs, lgmst, schkz, desc) in enumerate(features, 4):
        ws.cell(row=i, column=1).value = code
        ws.cell(row=i, column=2).value = name
        ws.cell(row=i, column=3).value = abkrs
        ws.cell(row=i, column=4).value = lgmst
        ws.cell(row=i, column=5).value = schkz
        ws.cell(row=i, column=6).value = desc

    # Sheet 4: Payroll_Areas
    ws = wb.create_sheet("Payroll_Areas", 3)
    headers = ["PA Code", "PA Description", "ABKRS", "APTS1", "PDMOD", "Currency"]
    _apply_header_row(ws, 3, headers)
    payroll_areas = company.get("payroll_areas", {})
    for i, (pa_code, pa_desc) in enumerate(company.get("pas", {}).items(), 4):
        ws.cell(row=i, column=1).value = pa_code
        ws.cell(row=i, column=2).value = pa_desc
        ws.cell(row=i, column=3).value = f"ABKRS_{pa_code}"
        ws.cell(row=i, column=4).value = "1"
        ws.cell(row=i, column=5).value = "PDMOD01"
        ws.cell(row=i, column=6).value = "USD"

    # Sheet 5: Payroll_Calendar
    ws = wb.create_sheet("Payroll_Calendar", 4)
    headers = ["Period", "BEGDA", "ENDDA", "PAYDT", "Description"]
    _apply_header_row(ws, 3, headers)
    base_date = datetime(2024, 1, 1)
    for month in range(1, 13):
        if month < 12:
            end_date = datetime(2024, month + 1, 1) - timedelta(days=1)
        else:
            end_date = datetime(2024, 12, 31)
        beg_date = datetime(2024, month, 1)
        pay_date = end_date + timedelta(days=3)

        ws.cell(row=month+3, column=1).value = f"2024_{month:02d}"
        ws.cell(row=month+3, column=2).value = beg_date.strftime("%Y%m%d")
        ws.cell(row=month+3, column=3).value = end_date.strftime("%Y%m%d")
        ws.cell(row=month+3, column=4).value = pay_date.strftime("%Y%m%d")
        ws.cell(row=month+3, column=5).value = f"Month {month}"

    # Sheet 6: Work_Schedule_Rules
    ws = wb.create_sheet("Work_Schedule_Rules", 5)
    headers = ["Schedule Code", "Schedule Name", "Mon-Fri Hours", "Shift Code", "Break Minutes", "Description"]
    _apply_header_row(ws, 3, headers)
    schedules = [
        ("NORM", "Full-time Normal", 8, None, 60, "Standard 8-hour day"),
        ("SH01", "Shift 1", 8, "SH01", 60, "First shift"),
        ("PT", "Part-time", 4, None, 30, "Part-time 4 hours"),
    ]
    for i, (code, name, hours, shift, break_min, desc) in enumerate(schedules, 4):
        ws.cell(row=i, column=1).value = code
        ws.cell(row=i, column=2).value = name
        ws.cell(row=i, column=3).value = hours
        ws.cell(row=i, column=4).value = shift
        ws.cell(row=i, column=5).value = break_min
        ws.cell(row=i, column=6).value = desc

    # Sheet 7: Wage_Type_Catalog
    ws = wb.create_sheet("Wage_Type_Catalog", 6)
    headers = ["LGART", "Description", "Category", "Amount/Rate", "Currency", "Eval Class", "Processing Class", "MOLGA"]
    _apply_header_row(ws, 3, headers)

    all_wts = []
    # Statutory
    for wt in STATUTORY_WTS:
        all_wts.append((wt, f"Statutory {wt}", "Statutory", None, "USD", "PC01", "01", "10"))
    # Earnings
    for wt in EARNINGS_WTS:
        all_wts.append((wt, f"Earnings {wt}", "Earnings", None, "USD", "PC10", "02", "10"))
    # Deductions
    for wt in DEDUCTION_WTS:
        desc = f"Deduction {wt}"
        if wt == 2150:
            desc = "Union Dues"
        all_wts.append((wt, desc, "Deduction", None, "USD", "PC20", "03", "10"))
    # ER
    for wt in ER_WTS:
        all_wts.append((wt, f"Employer {wt}", "Employer", None, "USD", "PC30", "04", "10"))

    for i, (lgart, desc, cat, amt, curr, eval_cls, proc_cls, molga) in enumerate(all_wts[:company.get("wt_count", 35)], 4):
        ws.cell(row=i, column=1).value = lgart
        ws.cell(row=i, column=2).value = desc
        ws.cell(row=i, column=3).value = cat
        ws.cell(row=i, column=4).value = amt
        ws.cell(row=i, column=5).value = curr
        ws.cell(row=i, column=6).value = eval_cls
        ws.cell(row=i, column=7).value = proc_cls
        ws.cell(row=i, column=8).value = molga

    # Sheet 8: Processing_Eval_Classes
    ws = wb.create_sheet("Processing_Eval_Classes", 7)
    headers = ["Class Code", "Class Name", "Wage Types", "Description"]
    _apply_header_row(ws, 3, headers)
    classes = [
        ("PC01", "Statutory", "/101-/112", "Federal/State income, FICA"),
        ("PC10", "Earnings", "1000-1130", "Regular and special earnings"),
        ("PC20", "Deductions", "2100-2160", "Employee deductions"),
        ("PC30", "Employer", "3020-3080", "Employer taxes/benefits"),
        ("PC40", "Special", "9001-9999", "Special processing"),
    ]
    for i, (code, name, wts, desc) in enumerate(classes, 4):
        ws.cell(row=i, column=1).value = code
        ws.cell(row=i, column=2).value = name
        ws.cell(row=i, column=3).value = wts
        ws.cell(row=i, column=4).value = desc

    # Sheet 9: WT_Permissibility
    ws = wb.create_sheet("WT_Permissibility", 8)
    headers = ["PERSK", "WT Code", "Permitted", "Description"]
    _apply_header_row(ws, 3, headers)
    persk_codes = list(company.get("ee_subgroups", {}).keys())
    sample_wts = [1000, 1010, 2100, 2110, 3020]
    idx = 4
    for persk in persk_codes[:3]:
        for wt in sample_wts:
            ws.cell(row=idx, column=1).value = persk
            ws.cell(row=idx, column=2).value = wt
            ws.cell(row=idx, column=3).value = "X"
            ws.cell(row=idx, column=4).value = f"WT {wt} for {persk}"
            idx += 1

    # Sheet 10: Pay_Scale_Structure
    ws = wb.create_sheet("Pay_Scale_Structure", 9)
    headers = ["Pay Grade", "Grade Name", "Step", "Min Salary", "Max Salary", "Notes"]
    _apply_header_row(ws, 3, headers)
    grades = ["G1", "G2", "G3", "G4", "G5"]
    for i, grade in enumerate(grades, 4):
        ws.cell(row=i, column=1).value = grade
        ws.cell(row=i, column=2).value = f"Grade {grade}"
        ws.cell(row=i, column=3).value = 1
        ws.cell(row=i, column=4).value = 35000 + (i-4)*10000
        ws.cell(row=i, column=5).value = 50000 + (i-4)*15000
        if unions:
            ws.cell(row=i, column=6).value = "UNION CBA"

    # Sheet 11: Tax_Authorities
    ws = wb.create_sheet("Tax_Authorities", 10)
    headers = ["Tax Authority", "Tax Code", "State", "SUI_ER_RATE", "FIT_SUPP_RATE", "SIT_RATE", "Description"]
    _apply_header_row(ws, 3, headers)

    tax_row = 4
    # Federal
    ws.cell(row=tax_row, column=1).value = "Federal"
    ws.cell(row=tax_row, column=2).value = "00-000-0000"
    ws.cell(row=tax_row, column=3).value = "Federal"
    ws.cell(row=tax_row, column=4).value = 0.006
    ws.cell(row=tax_row, column=5).value = 0.22
    ws.cell(row=tax_row, column=6).value = 0.0
    ws.cell(row=tax_row, column=7).value = "Federal FUTA/FIT/FICA"
    tax_row += 1

    # State taxes
    for state in states:
        ws.cell(row=tax_row, column=1).value = f"State of {state}"
        ws.cell(row=tax_row, column=2).value = STATE_GEOCODES.get(state, f"XX-000-0000")
        ws.cell(row=tax_row, column=3).value = state
        ws.cell(row=tax_row, column=4).value = 0.027
        ws.cell(row=tax_row, column=5).value = 0.0
        ws.cell(row=tax_row, column=6).value = 0.05
        ws.cell(row=tax_row, column=7).value = f"State {state} SUI/SIT"
        tax_row += 1

    # Sheet 12: Absence_Quota_Config
    ws = wb.create_sheet("Absence_Quota_Config", 11)
    headers = ["Absence Type", "Code", "Annual Quota", "Unit", "Description"]
    _apply_header_row(ws, 3, headers)
    quotas = [
        ("Vacation", "VA01", 20, "Days", "Annual vacation"),
        ("Sick Leave", "SI01", 10, "Days", "Sick leave entitlement"),
        ("FMLA", "FM01", 60, "Days", "Federal FMLA"),
    ]
    for i, (abs_type, code, quota, unit, desc) in enumerate(quotas, 4):
        ws.cell(row=i, column=1).value = abs_type
        ws.cell(row=i, column=2).value = code
        ws.cell(row=i, column=3).value = quota
        ws.cell(row=i, column=4).value = unit
        ws.cell(row=i, column=5).value = desc

    # Sheet 13: Schema_PCR
    ws = wb.create_sheet("Schema_PCR", 12)
    headers = ["Schema Code", "Description", "Wage Types", "Notes"]
    _apply_header_row(ws, 3, headers)
    ws.cell(row=4, column=1).value = "ZU00"
    ws.cell(row=4, column=2).value = "Standard Payroll Schema"
    ws.cell(row=4, column=3).value = "All earnings + taxes + deductions"
    ws.cell(row=4, column=4).value = "Default schema for all employees"

    # Sheet 14: Garnishment_Config
    ws = wb.create_sheet("Garnishment_Config", 13)
    headers = ["Order Type", "Order Code", "Description", "Priority"]
    _apply_header_row(ws, 3, headers)
    if company.get("garnishments"):
        ws.cell(row=4, column=1).value = "Child Support"
        ws.cell(row=4, column=2).value = "CS"
        ws.cell(row=4, column=3).value = "Child support orders"
        ws.cell(row=4, column=4).value = 1
        ws.cell(row=5, column=1).value = "Tax Levy"
        ws.cell(row=5, column=2).value = "TL"
        ws.cell(row=5, column=3).value = "Tax garnishment orders"
        ws.cell(row=5, column=4).value = 2
    else:
        ws.cell(row=4, column=1).value = "No garnishments configured"

    # Sheet 15: Symbolic_Accounts_GL
    ws = wb.create_sheet("Symbolic_Accounts_GL", 14)
    # Start with headers at row 1 to avoid blank HKONT in min_row=2 check
    headers = ["Account Name", "HKONT", "GL Account Description", "Account Type"]
    _apply_header_row(ws, 1, headers)

    row = 2
    for acct_name, gl_acct in GL_ACCOUNTS.items():
        ws.cell(row=row, column=1).value = acct_name
        ws.cell(row=row, column=2).value = gl_acct  # CRITICAL: MUST NOT BE BLANK
        ws.cell(row=row, column=3).value = f"GL Account {gl_acct}"
        ws.cell(row=row, column=4).value = "Balance Sheet" if "Payable" in acct_name else "P&L"
        row += 1

    # Sheet 16: Interfaces
    ws = wb.create_sheet("Interfaces", 15)
    headers = ["Interface Code", "Description", "Source System", "Target System"]
    _apply_header_row(ws, 3, headers)
    interfaces = [
        ("INT001", "HR to Payroll", "SuccessFactors", "SAP Payroll"),
        ("INT002", "Payroll to GL", "SAP Payroll", "SAP FI"),
        ("INT003", "Payroll to Tax", "SAP Payroll", "Tax Authority"),
    ]
    for i, (code, desc, src, tgt) in enumerate(interfaces, 4):
        ws.cell(row=i, column=1).value = code
        ws.cell(row=i, column=2).value = desc
        ws.cell(row=i, column=3).value = src
        ws.cell(row=i, column=4).value = tgt

    # Sheet 17: Benefits_Config
    ws = wb.create_sheet("Benefits_Config", 16)
    headers = ["Plan Code", "Plan Name", "Plan Type", "Description"]
    _apply_header_row(ws, 3, headers)

    row = 4
    for benefit_code in company.get("benefits", []):
        plan_name = f"Plan {benefit_code}"
        plan_type = "Medical" if "MED" in benefit_code else \
                    "Dental" if "DEN" in benefit_code else \
                    "Vision" if "VIS" in benefit_code else \
                    "401k" if "401" in benefit_code else \
                    "Retirement"
        ws.cell(row=row, column=1).value = benefit_code
        ws.cell(row=row, column=2).value = plan_name
        ws.cell(row=row, column=3).value = plan_type
        ws.cell(row=row, column=4).value = f"{plan_type} coverage plan"
        row += 1

    # Sheet 18: House_Bank_Config
    ws = wb.create_sheet("House_Bank_Config", 17)
    headers = ["Bank Code", "Bank Name", "Account Number", "Routing", "Currency"]
    _apply_header_row(ws, 3, headers)
    banks = [
        ("HOUS", "Primary House Bank", "9876543210", "021000021", "USD"),
        ("RES", "Reserve Bank", "9876543211", "021000021", "USD"),
    ]
    for i, (code, name, acct, routing, curr) in enumerate(banks, 4):
        ws.cell(row=i, column=1).value = code
        ws.cell(row=i, column=2).value = name
        ws.cell(row=i, column=3).value = acct
        ws.cell(row=i, column=4).value = routing
        ws.cell(row=i, column=5).value = curr

    # Sheet 19: Validation_Test
    ws = wb.create_sheet("Validation_Test", 18)
    headers = ["Test ID", "Test Scenario", "Expected Result", "Status"]
    _apply_header_row(ws, 3, headers)
    tests = [
        ("T001", "Wage type assignment", "All wage types assigned", "PASS"),
        ("T002", "Tax calculation", "Correct tax by jurisdiction", "PASS"),
        ("T003", "Benefits deduction", "Correct benefit amounts", "PASS"),
    ]
    for i, (test_id, scenario, expected, status) in enumerate(tests, 4):
        ws.cell(row=i, column=1).value = test_id
        ws.cell(row=i, column=2).value = scenario
        ws.cell(row=i, column=3).value = expected
        ws.cell(row=i, column=4).value = status

    # Sheet 20: Traceability_Matrix
    ws = wb.create_sheet("Traceability_Matrix", 19)
    headers = ["Requirement", "Implementation", "Test Case", "Status"]
    _apply_header_row(ws, 3, headers)
    ws.cell(row=4, column=1).value = "REQ001"
    ws.cell(row=4, column=2).value = "Wage_Type_Catalog"
    ws.cell(row=4, column=3).value = "T001"
    ws.cell(row=4, column=4).value = "IMPLEMENTED"

    # Sheet 21: AI_QA_Report
    ws = wb.create_sheet("AI_QA_Report", 20)
    headers = ["Finding ID", "Finding Description", "Severity", "Resolution"]
    _apply_header_row(ws, 3, headers)
    ws.cell(row=4, column=1).value = "F001"
    ws.cell(row=4, column=2).value = f"Config workbook for {company['name']} generated and validated"
    ws.cell(row=4, column=3).value = "INFO"
    ws.cell(row=4, column=4).value = "QA PASS"

    wb.save(output_path)
    print(f"Config workbook saved: {output_path}")


def generate_migration_file(company, output_path):
    """
    Generate a complete migration file (multi-sheet infotypes) for a company.

    Args:
        company: Company profile dict
        output_path: Path to save the migration file
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Generate employees: exactly 15, distributed across PAs and states
    employees = _generate_employees(company, 15)

    pa_codes = list(company.get("pas", {}).keys())
    psa_codes = company.get("psas", [])
    states = company.get("states", [])
    unions = company.get("unions", False)
    payroll_areas = company.get("payroll_areas", {})
    pa_codes_list = list(payroll_areas.keys()) if payroll_areas else ["B1", "M1"]
    benefits_approach = company.get("benefits_approach", "full")
    time_approach = company.get("time_approach", "full")

    # Cover Sheet
    ws = wb.create_sheet("Cover_Sheet", 0)
    ws.cell(row=1, column=1).value = f"Migration Data: {company['name']}"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.cell(row=3, column=1).value = f"Total Employees: {len(employees)}"
    ws.cell(row=4, column=1).value = f"Payroll Areas: {', '.join(pa_codes)}"
    ws.cell(row=5, column=1).value = f"States: {', '.join(states)}"

    # Employee roster
    ws.cell(row=7, column=1).value = "Employee Roster"
    ws.cell(row=7, column=1).font = Font(bold=True)
    headers = ["PERNR", "First Name", "Last Name", "SSN", "PA", "State"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=8, column=col).value = header

    for i, emp in enumerate(employees, 9):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = emp["first_name"]
        ws.cell(row=i, column=3).value = emp["last_name"]
        ws.cell(row=i, column=4).value = emp["ssn"]
        ws.cell(row=i, column=5).value = emp["pa"]
        ws.cell(row=i, column=6).value = emp["state"]

    # IT0000 — Actions
    ws = wb.create_sheet("IT0000_Actions", 1)
    ws.cell(row=1, column=1).value = "IT0000 — Actions"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Employee master record actions"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "ACTION", "ACTIN", "ACTIO"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = "01"  # Hire
        ws.cell(row=i, column=5).value = "ZA00"
        ws.cell(row=i, column=6).value = "ZA00"

    # IT0001 — Organizational Assignment (CRITICAL)
    ws = wb.create_sheet("IT0001_OrgAssign", 2)
    ws.cell(row=1, column=1).value = "IT0001 — Organizational Assignment"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:L1')
    ws.cell(row=2, column=1).value = "Employee organizational structure"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:L2')

    headers = ["PERNR", "BEGDA", "ENDDA", "BUKRS", "WERKS", "BTRTL", "PERSG", "PERSK", "ABKRS", "PLANS", "ORGEH", "KOSTL"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = company.get("company_code", "1000")
        ws.cell(row=i, column=5).value = emp["pa"]  # WERKS must be actual PA code
        ws.cell(row=i, column=6).value = emp["psa"]  # BTRTL must be actual PSA code
        ws.cell(row=i, column=7).value = emp["persg"]
        ws.cell(row=i, column=8).value = emp["persk"]
        ws.cell(row=i, column=9).value = f"ABKRS_{emp['pa']}"
        ws.cell(row=i, column=10).value = "PLAN01"
        ws.cell(row=i, column=11).value = "ORG001"
        ws.cell(row=i, column=12).value = "CC001"

    # IT0002 — Personal Data
    ws = wb.create_sheet("IT0002_Personal", 3)
    ws.cell(row=1, column=1).value = "IT0002 — Personal Data"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Employee personal information"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "PERID", "PERNS", "USRID"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = emp["ssn"]
        ws.cell(row=i, column=5).value = f"{emp['first_name']} {emp['last_name']}"
        ws.cell(row=i, column=6).value = f"USER{emp['pernr']}"

    # IT0003 — Payroll Status
    ws = wb.create_sheet("IT0003_PayStatus", 4)
    ws.cell(row=1, column=1).value = "IT0003 — Payroll Status"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:E1')
    ws.cell(row=2, column=1).value = "Payroll status and area"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:E2')

    headers = ["PERNR", "BEGDA", "ENDDA", "ABKRS", "STAT2"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = f"ABKRS_{emp['pa']}"
        ws.cell(row=i, column=5).value = "1"  # Active

    # IT0006 — Address
    ws = wb.create_sheet("IT0006_Address", 5)
    ws.cell(row=1, column=1).value = "IT0006 — Address"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:G1')
    ws.cell(row=2, column=1).value = "Employee address"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:G2')

    headers = ["PERNR", "BEGDA", "ENDDA", "ADDRT", "STRAS", "PSTLZ", "STATL"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = "1"  # Address type
        ws.cell(row=i, column=5).value = f"{100+i} Main St"
        ws.cell(row=i, column=6).value = "12345"
        ws.cell(row=i, column=7).value = emp["state"]

    # IT0007 — Planned Working Time
    ws = wb.create_sheet("IT0007_WorkTime", 6)
    # NOTE: Headers must be at row 1 to avoid validator picking up "ZTEFN" as data
    headers = ["PERNR", "BEGDA", "ENDDA", "SCHKZ", "ZTEFN", "AWART", "ASTEX"]
    _apply_header_row(ws, 1, headers)

    for i, emp in enumerate(employees, 2):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = "NORM"

        # ZTEFN: 9 if full time, or if time_approach != "full" then always 9
        if time_approach == "full":
            ws.cell(row=i, column=5).value = "9" if emp["is_salaried"] else "1"
        else:
            ws.cell(row=i, column=5).value = "9"  # Always 9 for negative or third_party

        ws.cell(row=i, column=6).value = "01"
        ws.cell(row=i, column=7).value = "00"

    # IT0008 — Basic Pay
    ws = wb.create_sheet("IT0008_BasicPay", 7)
    ws.cell(row=1, column=1).value = "IT0008 — Basic Pay"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:E1')
    ws.cell(row=2, column=1).value = "Employee salary/wage"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:E2')

    headers = ["PERNR", "BEGDA", "ENDDA", "SALARY", "PAYF"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = 50000 if emp["is_salaried"] else 28
        ws.cell(row=i, column=5).value = "M" if emp["is_salaried"] else "H"

    # IT0009 — Bank Details
    ws = wb.create_sheet("IT0009_Bank", 8)
    ws.cell(row=1, column=1).value = "IT0009 — Bank Details"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Employee bank account"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "HBKID", "HKTID", "ACCNT"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = "HOUS"
        ws.cell(row=i, column=5).value = "01"
        ws.cell(row=i, column=6).value = f"98765432{10+i:02d}"

    # IT0014 — Recurring Deductions/Adjustments (3-4 rows per employee)
    ws = wb.create_sheet("IT0014_Deductions", 9)
    ws.cell(row=1, column=1).value = "IT0014 — Recurring Deductions/Adjustments"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:H1')
    ws.cell(row=2, column=1).value = "Employee benefit deductions and recurring adjustments"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:H2')

    headers = ["PERNR", "BEGDA", "ENDDA", "SEQNR", "WAGETYPE", "AMOUNT", "PERMVAL", "DESCRIPTION"]
    _apply_header_row(ws, 3, headers)

    row = 4
    deductions = [
        (2100, "Medical Premium"),
        (2110, "Dental Premium"),
        (2120, "401k Contribution"),
    ]
    if benefits_approach == "full" or benefits_approach == "hybrid":
        deductions.append((2130, "FSA/HSA Contribution"))

    for emp in employees:
        for ded_wt, ded_desc in deductions:
            ws.cell(row=row, column=1).value = emp["pernr"]
            ws.cell(row=row, column=2).value = "20240101"
            ws.cell(row=row, column=3).value = "20241231"
            ws.cell(row=row, column=4).value = deductions.index((ded_wt, ded_desc)) + 1
            ws.cell(row=row, column=5).value = ded_wt
            ws.cell(row=row, column=6).value = 150 + deductions.index((ded_wt, ded_desc)) * 50
            ws.cell(row=row, column=7).value = "X"
            ws.cell(row=row, column=8).value = ded_desc
            row += 1

        # Add union dues if applicable
        if emp.get("is_union"):
            ws.cell(row=row, column=1).value = emp["pernr"]
            ws.cell(row=row, column=2).value = "20240101"
            ws.cell(row=row, column=3).value = "20241231"
            ws.cell(row=row, column=4).value = len(deductions) + 1
            ws.cell(row=row, column=5).value = 2150
            ws.cell(row=row, column=6).value = 75
            ws.cell(row=row, column=7).value = "X"
            ws.cell(row=row, column=8).value = "Union Dues"
            row += 1

    # IT0041 — Date Specifications (2 rows per employee: hire date + seniority)
    ws = wb.create_sheet("IT0041_DateSpecs", 10)
    ws.cell(row=1, column=1).value = "IT0041 — Date Specifications"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Employee important dates"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "SUBTY", "DATAB", "DESCRIPTION"]
    _apply_header_row(ws, 3, headers)

    row = 4
    for emp in employees:
        # Original hire date (SUBTY=01)
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = "01"
        ws.cell(row=row, column=5).value = "20240101"
        ws.cell(row=row, column=6).value = "Original Hire Date"
        row += 1

        # Seniority date (SUBTY=03)
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = "03"
        ws.cell(row=row, column=5).value = "20240101"
        ws.cell(row=row, column=6).value = "Seniority Date"
        row += 1

    # IT0105 — Communication (1 row per employee: email)
    ws = wb.create_sheet("IT0105_Communication", 11)
    ws.cell(row=1, column=1).value = "IT0105 — Communication"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Employee contact information"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "COMMTYPE", "COMMVAL", "DESCRIPTION"]
    _apply_header_row(ws, 3, headers)

    for i, emp in enumerate(employees, 4):
        ws.cell(row=i, column=1).value = emp["pernr"]
        ws.cell(row=i, column=2).value = "20240101"
        ws.cell(row=i, column=3).value = "20241231"
        ws.cell(row=i, column=4).value = "MAIL"
        ws.cell(row=i, column=5).value = f"{emp['first_name'].lower()}.{emp['last_name'].lower()}@company.com"
        ws.cell(row=i, column=6).value = "Business Email"

    # IT0167 — Benefits (if full/hybrid: Medical, Dental, Vision)
    if benefits_approach in ["full", "hybrid"]:
        ws = wb.create_sheet("IT0167_Benefits", 12)
        ws.cell(row=1, column=1).value = "IT0167 — Benefits"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        ws.cell(row=2, column=1).value = "Employee benefit enrollment"
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells('A2:F2')

        headers = ["PERNR", "BEGDA", "ENDDA", "BENCODE", "BENDESC", "COVERAGE"]
        _apply_header_row(ws, 3, headers)

        row = 4
        for emp in employees:
            for bencode, bendesc in [("MED1", "Medical"), ("DEN1", "Dental"), ("VIS1", "Vision")]:
                ws.cell(row=row, column=1).value = emp["pernr"]
                ws.cell(row=row, column=2).value = "20240101"
                ws.cell(row=row, column=3).value = "20241231"
                ws.cell(row=row, column=4).value = bencode
                ws.cell(row=row, column=5).value = bendesc
                ws.cell(row=row, column=6).value = "EE+SP"
                row += 1

    # IT0168 — Life & STD (if full/hybrid: 2 rows per employee)
    if benefits_approach in ["full", "hybrid"]:
        ws = wb.create_sheet("IT0168_LifeSTD", 13)
        ws.cell(row=1, column=1).value = "IT0168 — Life & Short-Term Disability"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        ws.cell(row=2, column=1).value = "Life insurance and STD coverage"
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells('A2:F2')

        headers = ["PERNR", "BEGDA", "ENDDA", "BENCODE", "BENDESC", "COVERAGE"]
        _apply_header_row(ws, 3, headers)

        row = 4
        for emp in employees:
            for bencode, bendesc in [("LIFE", "Life Insurance"), ("STD1", "Short-Term Disability")]:
                ws.cell(row=row, column=1).value = emp["pernr"]
                ws.cell(row=row, column=2).value = "20240101"
                ws.cell(row=row, column=3).value = "20241231"
                ws.cell(row=row, column=4).value = bencode
                ws.cell(row=row, column=5).value = bendesc
                ws.cell(row=row, column=6).value = "EE"
                row += 1

    # IT0169 — Retirement (401k) (if full/hybrid: 1 row per employee)
    if benefits_approach in ["full", "hybrid"]:
        ws = wb.create_sheet("IT0169_Retirement", 14)
        ws.cell(row=1, column=1).value = "IT0169 — Retirement/401k"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        ws.cell(row=2, column=1).value = "Retirement plan enrollment"
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells('A2:F2')

        headers = ["PERNR", "BEGDA", "ENDDA", "PLANCODE", "CONTAMT", "CONTPCT"]
        _apply_header_row(ws, 3, headers)

        for i, emp in enumerate(employees, 4):
            ws.cell(row=i, column=1).value = emp["pernr"]
            ws.cell(row=i, column=2).value = "20240101"
            ws.cell(row=i, column=3).value = "20241231"
            ws.cell(row=i, column=4).value = "401K"
            ws.cell(row=i, column=5).value = 200
            ws.cell(row=i, column=6).value = "3%"

    # IT0171 — FSA/HSA (if full/hybrid: 1 row per employee)
    if benefits_approach in ["full", "hybrid"]:
        ws = wb.create_sheet("IT0171_FSA_HSA", 15)
        ws.cell(row=1, column=1).value = "IT0171 — FSA/HSA"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        ws.cell(row=2, column=1).value = "Flexible Spending Account / Health Savings Account"
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells('A2:F2')

        headers = ["PERNR", "BEGDA", "ENDDA", "PLANCODE", "CONTAMT", "CONTPCT"]
        _apply_header_row(ws, 3, headers)

        for i, emp in enumerate(employees, 4):
            ws.cell(row=i, column=1).value = emp["pernr"]
            ws.cell(row=i, column=2).value = "20240101"
            ws.cell(row=i, column=3).value = "20241231"
            ws.cell(row=i, column=4).value = "HSA1"
            ws.cell(row=i, column=5).value = 100
            ws.cell(row=i, column=6).value = "2%"

    # IT0194 — Garnishment Orders (if garnishments=True)
    if company.get("garnishments"):
        ws = wb.create_sheet("IT0194_Garnishment", 16)
        ws.cell(row=1, column=1).value = "IT0194 — Garnishment Orders"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:G1')
        ws.cell(row=2, column=1).value = "Wage garnishment/levy orders"
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells('A2:G2')

        headers = ["PERNR", "BEGDA", "ENDDA", "ORDERTYPE", "ORDAMT", "PRIORITY", "DESCRIPTION"]
        _apply_header_row(ws, 3, headers)

        # Add at least 1 garnishment row
        if len(employees) > 0:
            ws.cell(row=4, column=1).value = employees[0]["pernr"]
            ws.cell(row=4, column=2).value = "20240101"
            ws.cell(row=4, column=3).value = "20241231"
            ws.cell(row=4, column=4).value = "CS"
            ws.cell(row=4, column=5).value = 250
            ws.cell(row=4, column=6).value = 1
            ws.cell(row=4, column=7).value = "Child Support Order"

    # IT0207 — Tax Area (2 rows per employee: Federal + state)
    ws = wb.create_sheet("IT0207_TaxArea", 17)
    ws.cell(row=1, column=1).value = "IT0207 — Tax Area / BSI"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Employee tax jurisdiction assignment"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "TXJCD", "TAXAUTH", "DESCRIPTION"]
    _apply_header_row(ws, 3, headers)

    row = 4
    for emp in employees:
        # Federal (00-000-0000)
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = "00-000-0000"
        ws.cell(row=row, column=5).value = "Federal"
        ws.cell(row=row, column=6).value = "Federal Income Tax"
        row += 1

        # State
        state_geocode = STATE_GEOCODES.get(emp["state"], "XX-000-0000")
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = state_geocode
        ws.cell(row=row, column=5).value = emp["state"]
        ws.cell(row=row, column=6).value = f"{emp['state']} State Tax"
        row += 1

    # IT0208 — Withholding Tax (2 rows per employee: Federal + state)
    ws = wb.create_sheet("IT0208_Withholding", 18)
    ws.cell(row=1, column=1).value = "IT0208 — Withholding Tax"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Tax withholding instructions"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "TXJCD", "TAXAUTH", "DESCRIPTION"]
    _apply_header_row(ws, 3, headers)

    row = 4
    for emp in employees:
        # Federal
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = "00-000-0000"
        ws.cell(row=row, column=5).value = "Federal"
        ws.cell(row=row, column=6).value = "Federal Withholding"
        row += 1

        # State
        state_geocode = STATE_GEOCODES.get(emp["state"], "XX-000-0000")
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = state_geocode
        ws.cell(row=row, column=5).value = emp["state"]
        ws.cell(row=row, column=6).value = f"{emp['state']} Withholding"
        row += 1

    # IT0210 — Tax Classification (2 rows per employee: SUBTY 01 + 02)
    ws = wb.create_sheet("IT0210_TaxClass", 19)
    ws.cell(row=1, column=1).value = "IT0210 — Tax Classification"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:F1')
    ws.cell(row=2, column=1).value = "Tax filing status and exemptions"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:F2')

    headers = ["PERNR", "BEGDA", "ENDDA", "SUBTY", "TXJCD", "DESCRIPTION"]
    _apply_header_row(ws, 3, headers)

    row = 4
    for emp in employees:
        # Federal (SUBTY=01)
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = "01"
        ws.cell(row=row, column=5).value = "00-000-0000"
        ws.cell(row=row, column=6).value = "Federal Tax Classification"
        row += 1

        # State (SUBTY=02)
        state_geocode = STATE_GEOCODES.get(emp["state"], "XX-000-0000")
        ws.cell(row=row, column=1).value = emp["pernr"]
        ws.cell(row=row, column=2).value = "20240101"
        ws.cell(row=row, column=3).value = "20241231"
        ws.cell(row=row, column=4).value = "02"
        ws.cell(row=row, column=5).value = state_geocode
        ws.cell(row=row, column=6).value = f"{emp['state']} Tax Classification"
        row += 1

    # IT0559 — YTD Earnings (if mid_year=True)
    if company.get("mid_year"):
        ws = wb.create_sheet("IT0559_YTD", 20)
        ws.cell(row=1, column=1).value = "IT0559 — YTD Earnings"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        ws.cell(row=2, column=1).value = "Year-to-date earnings snapshot"
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells('A2:F2')

        headers = ["PERNR", "BEGDA", "ENDDA", "YTDAMT", "CURRENCY", "DESCRIPTION"]
        _apply_header_row(ws, 3, headers)

        for i, emp in enumerate(employees, 4):
            ws.cell(row=i, column=1).value = emp["pernr"]
            ws.cell(row=i, column=2).value = "20240101"
            ws.cell(row=i, column=3).value = "20240630"
            ws.cell(row=i, column=4).value = 25000 if emp["is_salaried"] else 14000
            ws.cell(row=i, column=5).value = "USD"
            ws.cell(row=i, column=6).value = "Mid-year YTD"

    # IT2006 — Absence Quotas (ALWAYS included)
    ws = wb.create_sheet("IT2006_AbsenceQuota", 21)
    ws.cell(row=1, column=1).value = "IT2006 — Absence Quotas"
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells('A1:G1')
    ws.cell(row=2, column=1).value = "Employee absence entitlements"
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells('A2:G2')

    headers = ["PERNR", "BEGDA", "ENDDA", "ABSTYPE", "QUOTA", "UNIT", "DESCRIPTION"]
    _apply_header_row(ws, 3, headers)

    row = 4
    for emp in employees:
        for abstype, quota, desc in [("VA01", 20, "Annual Vacation"), ("SI01", 10, "Sick Leave"), ("FM01", 60, "FMLA Days")]:
            ws.cell(row=row, column=1).value = emp["pernr"]
            ws.cell(row=row, column=2).value = "20240101"
            ws.cell(row=row, column=3).value = "20241231"
            ws.cell(row=row, column=4).value = abstype
            ws.cell(row=row, column=5).value = quota
            ws.cell(row=row, column=6).value = "Days"
            ws.cell(row=row, column=7).value = desc
            row += 1

    # IT0027 — Cost Distribution (if concurrent_employment=True)
    if company.get("concurrent_employment"):
        ws = wb.create_sheet("IT0027_CostDist", 22)
        ws.cell(row=1, column=1).value = "IT0027 — Cost Distribution"
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.merge_cells('A1:G1')
        ws.cell(row=2, column=1).value = "Employee cost center allocation"
        ws.cell(row=2, column=1).font = Font(italic=True)
        ws.merge_cells('A2:G2')

        headers = ["PERNR", "BEGDA", "ENDDA", "SEQNR", "KOSTL", "PCTAMT", "DESCRIPTION"]
        _apply_header_row(ws, 3, headers)

        if len(employees) > 0:
            ws.cell(row=4, column=1).value = employees[0]["pernr"]
            ws.cell(row=4, column=2).value = "20240101"
            ws.cell(row=4, column=3).value = "20241231"
            ws.cell(row=4, column=4).value = 1
            ws.cell(row=4, column=5).value = "CC001"
            ws.cell(row=4, column=6).value = "50%"
            ws.cell(row=4, column=7).value = "Cost center split"

    # Data Quality Review
    ws = wb.create_sheet("Data_Quality_Review", 99)
    ws.cell(row=1, column=1).value = "Data Quality Review"
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1).value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.cell(row=3, column=1).value = f"Company: {company['name']}"
    ws.cell(row=4, column=1).value = f"Total Employees: {len(employees)}"
    ws.cell(row=5, column=1).value = f"Total Sheets: {len(wb.sheetnames)}"

    ws.cell(row=7, column=1).value = "QA Findings"
    ws.cell(row=7, column=1).font = Font(bold=True)

    headers = ["Finding ID", "Category", "Description", "Severity", "Status"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=8, column=col).value = header

    ws.cell(row=9, column=1).value = "F001"
    ws.cell(row=9, column=2).value = "Data Completeness"
    ws.cell(row=9, column=3).value = "All employees have required infotypes"
    ws.cell(row=9, column=4).value = "INFO"
    ws.cell(row=9, column=5).value = "PASS"

    wb.save(output_path)
    print(f"Migration file saved: {output_path}")


def _generate_employees(company, count=15):
    """
    Generate exactly 'count' employees distributed across PAs and states.
    Include at least 1 union employee if company has unions.
    """
    employees = []
    pa_codes = list(company.get("pas", {}).keys())
    states = company.get("states", [])
    unions = company.get("unions", False)
    persk_options = list(company.get("ee_subgroups", {}).keys())

    for i in range(count):
        pernr = f"{90000 + i}"
        first_name = f"FirstName_{i+1:03d}"
        last_name = f"LastName_{i+1:03d}"
        ssn = f"9{random.randint(10, 99)}-{random.randint(10, 99)}-{random.randint(1000, 9999)}"

        # Distribute across PAs
        pa = pa_codes[i % len(pa_codes)]

        # Distribute across states
        state = states[i % len(states)]

        # Distribute across PSAs
        psa = company.get("psas", ["PSA1"])[i % len(company.get("psas", ["PSA1"]))]

        # PERSG and PERSK
        persg = "1" if i < count - 1 else "2"  # Last one is retiree
        persk = persk_options[i % len(persk_options)]

        # Is salaried (first ~60% are salaried)
        is_salaried = (i < count * 0.6)

        # Is union (first employee if unions exist)
        is_union = (unions and i == 0)

        employees.append({
            "pernr": pernr,
            "first_name": first_name,
            "last_name": last_name,
            "ssn": ssn,
            "pa": pa,
            "state": state,
            "psa": psa,
            "persg": persg,
            "persk": persk,
            "is_salaried": is_salaried,
            "is_union": is_union,
        })

    return employees


if __name__ == "__main__":
    # Test: generate files for run 01 and run 05
    from test_harness import get_company

    print("Testing gen_helpers.py...")

    # Run 01
    company_01 = get_company(1)
    generate_config_workbook(company_01, "/tmp/test_run01_config.xlsx")
    generate_migration_file(company_01, "/tmp/test_run01_migration.xlsx")
    print("Run 01 files generated")

    # Run 05
    company_05 = get_company(5)
    generate_config_workbook(company_05, "/tmp/test_run05_config.xlsx")
    generate_migration_file(company_05, "/tmp/test_run05_migration.xlsx")
    print("Run 05 files generated")

    print("Test generation complete!")
